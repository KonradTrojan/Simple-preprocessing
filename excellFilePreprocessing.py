import openpyxl as op
import csv


def get_excel_sheet(path):

    wb_obj = op.load_workbook(path, data_only=True)

    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column  # it shows bad value, 15 instead of 5. After copying cells to other file it's ok

    data = []

    for i in range(1, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            row.append(cell_obj.value)
        data.append(row)
    return data


# convert xlsx file to csv
def xlsx_to_csv(from_path, to_path):

    with open(to_path, 'w', newline='') as f:
        write = csv.writer(f)
        write.writerows(get_excel_sheet(from_path))


if __name__ == '__main__':

    path = "files/data.xlsx"
    print(get_excel_sheet(path))
    xlsx_to_csv("files/data.xlsx", "new_files/new_data.csv")

